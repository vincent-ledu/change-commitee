#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Ajoute une heure aléatoire aux colonnes
- "Date de début planifiée"
- "Date de fin planifiée"

Et sauvegarde un nouveau fichier Excel avec des colonnes typées Date+Heure.

Usage simple:
  python add_random_times_to_dataset.py \
    --in cab_changes_5_weeks.xlsx \
    --out cab_change_data.xlsx

Dépendances: pandas, openpyxl (pour écrire en xlsx). Si openpyxl est disponible,
le script applique aussi un format d'affichage "dd/mm/yyyy hh:mm".
"""

from __future__ import annotations
import argparse
import random
from datetime import datetime, time
import pandas as pd


INPUT_COL_START = "Date de début planifiée"
INPUT_COL_END = "Date de fin planifiée"


def parse_fr_date_any(s: str) -> datetime | None:
    s = str(s or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except Exception:
            pass
    # tente une auto-détection pandas
    try:
        dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return dt.to_pydatetime()
    except Exception:
        pass
    return None


def with_random_time(d: datetime) -> datetime:
    h = random.randint(0, 23)
    m = random.randint(0, 59)
    return datetime.combine(d.date(), time(hour=h, minute=m))


def add_times(df: pd.DataFrame) -> pd.DataFrame:
    for col in (INPUT_COL_START, INPUT_COL_END):
        if col not in df.columns:
            raise ValueError(f"Colonne manquante: {col}")

    out = df.copy()

    def convert_col(colname: str) -> pd.Series:
        vals: list[datetime | None] = []
        for v in out[colname].tolist():
            base = parse_fr_date_any(v)
            if base is None:
                vals.append(pd.NaT)
            else:
                vals.append(with_random_time(base))
        return pd.to_datetime(pd.Series(vals), errors="coerce")

    out[INPUT_COL_START] = convert_col(INPUT_COL_START)
    out[INPUT_COL_END] = convert_col(INPUT_COL_END)
    return out


def write_excel_with_formats(df: pd.DataFrame, out_path: str) -> None:
    # Écrit via pandas/openpyxl, puis applique un format d'affichage si openpyxl est dispo.
    df.to_excel(out_path, index=False)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(out_path)
        ws = wb.active
        # applique le format aux colonnes correspondantes
        # recherche des index de colonnes (1-based)
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        fmt = "dd/mm/yyyy hh:mm"
        for col_name in (INPUT_COL_START, INPUT_COL_END):
            if col_name in header:
                idx = header.index(col_name) + 1
                for row in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
                    cell = row[0]
                    cell.number_format = fmt
        wb.save(out_path)
    except Exception:
        # Si openpyxl n'est pas dispo ou autre souci, on laisse le fichier tel quel
        pass


def main():
    ap = argparse.ArgumentParser(description="Ajoute une heure aléatoire aux colonnes de dates et écrit un nouvel Excel.")
    ap.add_argument("--in", dest="inp", default="cab_changes_5_weeks.xlsx", help="Fichier Excel source")
    ap.add_argument("--out", dest="out", default="cab_change_data.xlsx", help="Fichier Excel de sortie")
    args = ap.parse_args()

    df = pd.read_excel(args.inp, dtype=str)
    out = add_times(df)
    write_excel_with_formats(out, args.out)
    print(f"[OK] Écrit: {args.out}")


if __name__ == "__main__":
    main()

